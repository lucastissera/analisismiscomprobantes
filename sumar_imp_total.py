"""
Lee un archivo Excel (.xlsx) o CSV, ajusta signos según Tipo y suma columnas indicadas.
En .xlsx se prueba encabezado en fila 1 o fila 2 y se usa el que tenga todas las columnas requeridas.
En .csv los encabezados suelen estar en fila 1.
Las filas con Tipo = nota de crédito (por código numérico) se consideran en negativo.

Uso:
  python sumar_imp_total.py <ruta_al_archivo.xlsx> [hoja] [archivo_salida.xlsx]
"""

import io
import re
import sys
import csv
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# Columnas a evaluar: se les aplica signo por Tipo y multiplicación por Tipo Cambio
COLUMNAS_A_AJUSTAR = [
    "Neto Grav. IVA 0%",
    "IVA 2,5%",
    "Neto Grav. IVA 2,5%",
    "IVA 5%",
    "Neto Grav. IVA 5%",
    "IVA 10,5%",
    "Neto Grav. IVA 10,5%",
    "IVA 21%",
    "Neto Grav. IVA 21%",
    "IVA 27%",
    "Neto Grav. IVA 27%",
    "Neto Gravado Total",
    "Neto No Gravado",
    "Op. Exentas",
    "Otros Tributos",
    "Total IVA",
    "Imp. Total",
]

# Totales que se suman en la línea "Total" del resumen en pantalla (no IVA desglosado ni Imp. Total)
COLUMNAS_TOTAL_RESUMEN = [
    "Neto Grav. IVA 0%",
    "Neto Gravado Total",
    "Neto No Gravado",
    "Op. Exentas",
    "Otros Tributos",
    "Total IVA",
]

# Columnas de detalle (IVA por alícuota e Imp. Total): no entran en la fila "Total (resumen)"
COLUMNAS_DETALLE_SIN_RESUMEN = [
    c for c in COLUMNAS_A_AJUSTAR if c not in COLUMNAS_TOTAL_RESUMEN
]

# Notas de crédito: neto gravado e IVA por alícuota (2,5%–27%; resumen y reparto mensual)
COLUMNAS_NETO_NC_ALICUOTA = [
    "Neto Grav. IVA 2,5%",
    "Neto Grav. IVA 5%",
    "Neto Grav. IVA 10,5%",
    "Neto Grav. IVA 21%",
    "Neto Grav. IVA 27%",
]
COLUMNAS_IVA_NC_ALICUOTA = [
    "IVA 2,5%",
    "IVA 5%",
    "IVA 10,5%",
    "IVA 21%",
    "IVA 27%",
]

# Suma "Neto" en informe por proveedor/cliente (mismas columnas; valores ya ajustados)
COLUMNAS_NETO_INFORME_CONTRAPARTE = [
    "Neto Gravado Total",
    "Neto Grav. IVA 0%",
    "Neto No Gravado",
    "Op. Exentas",
]

# Códigos numéricos de la columna "Tipo" que se consideran nota de crédito (suma en negativo)
# Se matchea por número (sin ceros a la izquierda): "003", "3" y "03" son el mismo código
CODIGOS_NOTA_CREDITO = {
    3, 8, 13, 21, 38, 43, 44, 48, 53,
    110, 112, 113, 114, 203, 206, 208, 211, 213,
}

# Tipos B/C (y afines): Imp. Total → Neto Grav. IVA 0% (no Neto Gravado Total). Columna Tipo.
# Comprobantes recibidos: se aplica a todo este conjunto (export ARCA suele concentrar el total en Imp. Total).
CODIGOS_IMP_TOTAL_EN_NETO_IVA_0 = frozenset(
    (
        6, 7, 8, 9, 10, 11, 12, 13, 15, 16, 18, 19, 20, 21, 25, 26, 28, 29,
        40, 41, 42, 43, 44, 46, 47, 61, 64, 82, 83, 90, 91, 109, 110, 111,
        113, 114, 116, 117, 206, 207, 208, 211, 212, 213,
    )
)

# Solo clase C (AFIP Libro IVA / RG 4290 y afines; FCE MiPyME C = 211–213).
# Comprobantes emitidos: únicamente estos replican Imp. Total en Neto Grav. IVA 0% y anulan el resto
# de bases imponibles en esa fila; el resto de tipos usan las columnas del archivo (con signo NC).
CODIGOS_IMP_TOTAL_EN_NETO_IVA_0_SOLO_C = frozenset(
    (
        11,
        12,
        13,  # Factura / ND / NC C
        15,
        16,  # Recibo C, nota venta al contado C
        109,
        111,
        114,
        117,  # Tique C, tique factura C, tique NC/ND C
        211,
        212,
        213,  # FCE electrónica MiPyME C
    )
)

NOMBRES_MESES = [
    "enero",
    "febrero",
    "marzo",
    "abril",
    "mayo",
    "junio",
    "julio",
    "agosto",
    "septiembre",
    "octubre",
    "noviembre",
    "diciembre",
]

def limpiar_nombre_columna_bruto(nombre: str) -> str:
    """Corrige nombres de columna típicos (mojibake UTF-8 como latin-1, etc.)."""
    s = str(nombre).strip()
    s = s.replace("EmisiÃ³n", "Emisión")
    s = s.replace("Ã³", "ó").replace("Ã­", "í").replace("Ã¡", "á").replace("Ã©", "é")
    s = s.replace("Ãº", "ú").replace("Ã±", "ñ")
    return s


# Alias de columnas para soportar variaciones entre .xlsx y .csv (ARCA)
ALIAS_COLUMNAS = {
    "Fecha Emisión": [
        "Fecha Emisión",
        "Fecha de Emisión",
        "Fecha de emisión",
        "Fecha",
    ],
    "Tipo": ["Tipo", "Tipo de Comprobante"],
    "Tipo Cambio": ["Tipo Cambio"],
    "Neto Grav. IVA 0%": ["Neto Grav. IVA 0%", "Imp. Neto Gravado IVA 0%"],
    "IVA 2,5%": ["IVA 2,5%"],
    "Neto Grav. IVA 2,5%": ["Neto Grav. IVA 2,5%", "Imp. Neto Gravado IVA 2,5%"],
    "IVA 5%": ["IVA 5%"],
    "Neto Grav. IVA 5%": ["Neto Grav. IVA 5%", "Imp. Neto Gravado IVA 5%"],
    "IVA 10,5%": ["IVA 10,5%"],
    "Neto Grav. IVA 10,5%": ["Neto Grav. IVA 10,5%", "Imp. Neto Gravado IVA 10,5%"],
    "IVA 21%": ["IVA 21%"],
    "Neto Grav. IVA 21%": ["Neto Grav. IVA 21%", "Imp. Neto Gravado IVA 21%"],
    "IVA 27%": ["IVA 27%"],
    "Neto Grav. IVA 27%": ["Neto Grav. IVA 27%", "Imp. Neto Gravado IVA 27%"],
    "Neto Gravado Total": ["Neto Gravado Total", "Imp. Neto Gravado Total"],
    "Neto No Gravado": ["Neto No Gravado", "Imp. Neto No Gravado"],
    "Op. Exentas": ["Op. Exentas", "Imp. Op. Exentas"],
    "Otros Tributos": ["Otros Tributos"],
    "Total IVA": ["Total IVA"],
    "Imp. Total": ["Imp. Total"],
    "Denominación Emisor": [
        "Denominación Emisor",
        "Denominacion Emisor",
        "Razón Social Emisor",
    ],
    "Denominación Receptor": [
        "Denominación Receptor",
        "Denominacion Receptor",
        "Razón Social Receptor",
    ],
    "Nro. Doc. Emisor": [
        "Nro. Doc. Emisor",
        "Nro Doc. Emisor",
        "Nro Doc Emisor",
    ],
    "Nro. Doc. Receptor": [
        "Nro. Doc. Receptor",
        "Nro Doc. Receptor",
        "Nro Doc Receptor",
    ],
    "Número Hasta": ["Número Hasta", "Numero Hasta", "Nro. Hasta"],
    "Cód. Autorización": ["Cód. Autorización", "Cod. Autorización", "Cod Autorizacion"],
    "Tipo Doc. Emisor": ["Tipo Doc. Emisor", "Tipo Documento Emisor"],
    "Tipo Doc. Receptor": ["Tipo Doc. Receptor", "Tipo Documento Receptor"],
}


def parsear_numero_importe(val) -> float:
    """
    Convierte valores típicos de exportación ARCA/CSV argentino a float.
    Soporta coma decimal (10156,44), notación científica con coma (7,50154E+13)
    y separador de miles con punto (1.234,56).
    También acepta escalares numéricos de numpy/pandas ya leídos por read_csv.
    """
    if pd.isna(val):
        return float("nan")
    if isinstance(val, bool):
        return float("nan")
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return float(val)

    if isinstance(val, np.generic):
        return float(val)

    try:
        return float(val)
    except (TypeError, ValueError):
        pass

    s = str(val).strip()
    if not s or s in ("-", "–", "$"):
        return float("nan")

    # Notación científica con coma en la mantisa: 7,50154E+13
    if re.search(r"[Ee]", s):
        s = re.sub(
            r"^([-+]?\d+),(\d+[Ee][+-]?\d+)$",
            r"\1.\2",
            s,
            flags=re.IGNORECASE,
        )
        try:
            return float(s)
        except ValueError:
            return float("nan")

    s_clean = s.replace(" ", "")
    n_comma = s_clean.count(",")
    n_dot = s_clean.count(".")

    if n_comma == 1 and n_dot == 0:
        s_clean = s_clean.replace(",", ".")
    elif n_comma > 0 and n_dot > 0:
        if s_clean.rfind(",") > s_clean.rfind("."):
            s_clean = s_clean.replace(".", "").replace(",", ".")
        else:
            s_clean = s_clean.replace(",", "")
    elif n_comma > 1 and n_dot == 0:
        s_clean = s_clean.replace(",", ".", 1).replace(",", "", n_comma - 1)

    try:
        return float(s_clean)
    except ValueError:
        return float("nan")


def serie_a_float_importe(serie: pd.Series) -> pd.Series:
    """Convierte serie a float; prioriza dtype numérico y luego formato AR con coma."""
    if pd.api.types.is_numeric_dtype(serie):
        return pd.to_numeric(serie, errors="coerce")
    parsed = serie.map(parsear_numero_importe)
    # Si todo falló (p. ej. strings que pandas no interpretó), reintento vectorizado
    if parsed.notna().any() or serie.empty:
        return parsed
    str_serie = serie.astype(str).str.strip()
    str_serie = str_serie.replace("", pd.NA).replace("nan", pd.NA)
    coma_decimal = str_serie.str.contains(",", na=False) & ~str_serie.str.contains(
        r"[Ee]", na=False, regex=True
    )
    tmp = str_serie.copy()
    tmp.loc[coma_decimal] = tmp.loc[coma_decimal].str.replace(".", "", regex=False)
    tmp = tmp.str.replace(",", ".", regex=False)
    fallback = pd.to_numeric(tmp, errors="coerce")
    return fallback


def total_resumen_pantalla(totales: dict[str, float]) -> float:
    """Suma solo las columnas que deben aparecer en el total del resumen UI."""
    return float(sum(totales[c] for c in COLUMNAS_TOTAL_RESUMEN if c in totales))


def totales_resumen_por_periodo(
    totales_por_periodo: dict[str, dict[str, float]],
) -> dict[str, float]:
    """Total (resumen) por mes-año, misma regla que COLUMNAS_TOTAL_RESUMEN."""
    return {
        p: float(
            sum(
                totales_por_periodo[p][c]
                for c in COLUMNAS_TOTAL_RESUMEN
                if c in totales_por_periodo[p]
            )
        )
        for p in totales_por_periodo
    }


def periodos_orden_crono(*dicts) -> list[str]:
    """Unión de claves YYYY-MM presentes en diccionarios, orden cronológico."""
    keys: set[str] = set()
    for d in dicts:
        if d:
            keys |= set(d.keys())
    return sorted(keys)


def limpiar_argumento_ruta(valor: str) -> str:
    """Normaliza saltos de línea/tabulaciones accidentales en argumentos de ruta."""
    return valor.replace("\r", " ").replace("\n", " ").replace("\t", " ").strip()


def normalizar_columnas(df: pd.DataFrame) -> pd.DataFrame:
    """Renombra columnas conocidas a un nombre canónico interno."""
    df = df.copy()
    df.columns = [limpiar_nombre_columna_bruto(c) for c in df.columns]
    renombres = {}
    columnas_actuales = list(df.columns)
    for canonica, aliases in ALIAS_COLUMNAS.items():
        for alias in aliases:
            if alias in columnas_actuales:
                renombres[alias] = canonica
                break
    if renombres:
        df = df.rename(columns=renombres)
    # Evita df["Imp. Total"] como DataFrame si hubo nombres duplicados tras renombrar
    if df.columns.duplicated().any():
        df = df.loc[:, ~df.columns.duplicated()].copy()
    return df


def _columnas_requeridas_lectura() -> set[str]:
    return set(COLUMNAS_A_AJUSTAR + ["Tipo", "Tipo Cambio", "Fecha Emisión"])


def serie_codigo_tipo_comprobante(serie_tipo: pd.Series) -> pd.Series:
    """
    Código numérico AFIP en la columna Tipo (para isin con CODIGOS_*).
    - Excel a veces guarda el código como número (6, 11).
    - En texto: '6 - Factura B', '006 - ...', o guiones Unicode (– —) mal interpretados
      por split(' - '): se normalizan y, si hace falta, se toman dígitos iniciales.
    """
    s = serie_tipo
    if pd.api.types.is_numeric_dtype(s):
        return pd.to_numeric(s, errors="coerce")

    t = s.astype(str).str.strip()
    t = t.replace({"nan": pd.NA, "None": pd.NA, "<NA>": pd.NA})
    t = t.str.replace("\u2013", "-").str.replace("\u2014", "-")
    part0 = t.str.split(r"\s*-\s*", n=1, regex=True).str[0].str.strip()
    codigo = pd.to_numeric(part0, errors="coerce")
    vac = codigo.isna() & t.notna()
    if vac.any():
        dig = t.loc[vac].str.extract(r"^(\d+)", expand=False)
        codigo.loc[vac] = pd.to_numeric(dig, errors="coerce")
    return codigo


def _serie_fecha_emision_a_datetime(serie: pd.Series) -> pd.Series:
    """
    Convierte la columna Fecha Emisión a datetime64 (naïve o con TZ AR si aplica).
    Misma lógica para .xlsx (General/serial) y .csv (texto d/m/y, ISO, etc.).
    """
    serie = serie.reset_index(drop=True)
    n = len(serie)
    non_null = serie.notna()
    if n == 0 or not non_null.any():
        return pd.Series(pd.NaT, index=serie.index, dtype="datetime64[ns]")

    # --- 1) Serial Excel ---
    sn = pd.to_numeric(serie, errors="coerce")
    if non_null.any() and bool((sn.notna() == non_null).all()):
        vmin, vmax = float(sn[non_null].min()), float(sn[non_null].max())
        if vmin > 20000 and vmax < 80000:
            return pd.to_datetime(sn, unit="D", origin="1899-12-30", errors="coerce")

    # --- 2) datetime64 ---
    if pd.api.types.is_datetime64_any_dtype(serie.dtype):
        fechas = pd.to_datetime(serie, errors="coerce")
        if pd.api.types.is_datetime64tz_dtype(fechas):
            try:
                fechas = fechas.dt.tz_convert("America/Argentina/Buenos_Aires")
            except Exception:
                pass
        return fechas

    # --- 3) Texto ---
    s = serie.astype(str).str.strip()
    s = s.replace({"nan": pd.NA, "None": pd.NA, "<NA>": pd.NA, "NaT": pd.NA})
    s = s.str.replace(r"\s+\d{1,2}:\d{2}.*$", "", regex=True)
    # ISO primero (evita warning dayfirst con yyyy-mm-dd típico de CSV exportado)
    fechas = pd.to_datetime(s, format="%Y-%m-%d", errors="coerce")
    pend = fechas.isna() & s.notna()
    if pend.any():
        fechas.loc[pend] = pd.to_datetime(s.loc[pend], dayfirst=True, errors="coerce")
    pend = fechas.isna() & s.notna()
    if pend.any():
        fechas.loc[pend] = pd.to_datetime(
            s.loc[pend], format="%d/%m/%Y", errors="coerce"
        )
    pend = fechas.isna() & s.notna()
    if pend.any():
        fechas.loc[pend] = pd.to_datetime(
            s.loc[pend], format="%d-%m-%Y", errors="coerce"
        )
    pend = fechas.isna() & s.notna()
    if pend.any():
        try:
            fechas.loc[pend] = pd.to_datetime(
                s.loc[pend], dayfirst=True, errors="coerce", format="mixed"
            )
        except (TypeError, ValueError):
            pass

    return fechas


def _mes_fila_fecha_emision(
    df: pd.DataFrame, nombre_archivo: str | None
) -> pd.Series:
    """
    Mes calendario (1-12) por fila según Fecha Emisión (misma lógica .xlsx y .csv).
    """
    _ = nombre_archivo
    fechas = _serie_fecha_emision_a_datetime(df["Fecha Emisión"])
    mes = fechas.dt.month.astype(float)
    return mes.where(fechas.notna(), np.nan).reset_index(drop=True)


def _formatear_fecha_emision_salida_excel(df: pd.DataFrame) -> None:
    """
    Deja Fecha Emisión como texto dd/mm/yyyy en el DataFrame que se exporta a .xlsx,
    igual que suele verse al procesar .xlsx (evita que CSV salga como yyyy-mm-dd).
    No altera totales: debe llamarse después de _totales_anuales_y_por_mes.
    """
    if "Fecha Emisión" not in df.columns:
        return
    orig = df["Fecha Emisión"]
    fechas = _serie_fecha_emision_a_datetime(orig)
    texto = fechas.dt.strftime("%d/%m/%Y")
    df["Fecha Emisión"] = np.where(
        fechas.notna(),
        texto,
        np.where(orig.notna(), orig.astype(str), ""),
    )


def _totales_anuales_y_por_periodo(
    df_ajustado: pd.DataFrame,
    columnas: list[str],
    nombre_archivo: str | None,
) -> tuple[dict[str, float], dict[str, dict[str, float]]]:
    """
    Suma por columna (todas las filas) y acumulado por (año, mes) según Fecha Emisión.
    Claves de período: ``\"YYYY-MM\"`` (solo períodos con al menos un comprobante con fecha).
    """
    _ = nombre_archivo
    block = df_ajustado[columnas].apply(pd.to_numeric, errors="coerce").fillna(0.0)
    block_arr = block.to_numpy(dtype=np.float64, copy=False)
    resultado = {c: float(block_arr[:, j].sum()) for j, c in enumerate(columnas)}

    totales_por_periodo: dict[str, dict[str, float]] = {}

    fechas = _serie_fecha_emision_a_datetime(df_ajustado["Fecha Emisión"])
    años = fechas.dt.year
    meses = fechas.dt.month
    n = block_arr.shape[0]
    for pos in range(n):
        if pd.isna(fechas.iloc[pos]):
            continue
        y = int(años.iloc[pos])
        m = int(meses.iloc[pos])
        if m < 1 or m > 12:
            continue
        key = f"{y:04d}-{m:02d}"
        if key not in totales_por_periodo:
            totales_por_periodo[key] = {c: 0.0 for c in columnas}
        for j, c in enumerate(columnas):
            totales_por_periodo[key][c] += float(block_arr[pos, j])

    return resultado, totales_por_periodo


def totales_notas_credito_neto_e_iva_por_periodo(
    df_ajustado: pd.DataFrame,
    nombre_archivo: str | None,
) -> tuple[float, float, dict[str, float], dict[str, float]]:
    """
    Suma solo filas clasificadas como nota de crédito, sobre:
    - neto gravado 2,5% … 27% (sin 0%),
    - IVA 2,5% … 27%.

    Usa los valores ya en ``df_ajustado`` (misma regla B/C, signo NC y tipo de cambio
    que el resto). Período por Fecha de emisión, clave ``\"YYYY-MM\"``.
    Debe llamarse antes de formatear Fecha Emisión a texto si se desea; Tipo
    se sigue leyendo para el mask NC.
    """
    _ = nombre_archivo
    codigo_num = serie_codigo_tipo_comprobante(df_ajustado["Tipo"])
    es_nc = codigo_num.isin(CODIGOS_NOTA_CREDITO).to_numpy(dtype=bool)

    block_neto = df_ajustado[COLUMNAS_NETO_NC_ALICUOTA].apply(
        pd.to_numeric, errors="coerce"
    ).fillna(0.0)
    block_iva = df_ajustado[COLUMNAS_IVA_NC_ALICUOTA].apply(
        pd.to_numeric, errors="coerce"
    ).fillna(0.0)
    neto_row = block_neto.to_numpy(dtype=np.float64, copy=False).sum(axis=1)
    iva_row = block_iva.to_numpy(dtype=np.float64, copy=False).sum(axis=1)

    total_neto = float(neto_row[es_nc].sum()) if es_nc.any() else 0.0
    total_iva = float(iva_row[es_nc].sum()) if es_nc.any() else 0.0

    neto_por: dict[str, float] = defaultdict(float)
    iva_por: dict[str, float] = defaultdict(float)

    fechas = _serie_fecha_emision_a_datetime(df_ajustado["Fecha Emisión"])
    años = fechas.dt.year
    meses = fechas.dt.month
    n = int(block_neto.shape[0])
    if len(es_nc) > n:
        es_nc = es_nc[:n]
    elif len(es_nc) < n:
        es_nc = np.pad(es_nc, (0, n - len(es_nc)), constant_values=False)

    for pos in range(n):
        if not es_nc[pos]:
            continue
        if pd.isna(fechas.iloc[pos]):
            continue
        y = int(años.iloc[pos])
        m = int(meses.iloc[pos])
        if m < 1 or m > 12:
            continue
        key = f"{y:04d}-{m:02d}"
        neto_por[key] += float(neto_row[pos])
        iva_por[key] += float(iva_row[pos])

    return total_neto, total_iva, dict(neto_por), dict(iva_por)


def _normalizar_clave_cuit_doc(val) -> str:
    """
    Solo dígitos (DNI/CUIT). Si pandas/Excel leyó el documento como float
    (p. ej. 27318787949.0), ``str(val)`` sería ``27318787949.0`` y al quitar
    no-dígitos quedaría un 0 de más al final; se convierte antes a entero.
    """
    if val is None:
        return ""
    if isinstance(val, (int, np.integer)):
        if int(val) < 0:
            return ""
        return str(int(val))
    if isinstance(val, (float, np.floating)):
        f = float(val)
        if np.isnan(f) or not np.isfinite(f):
            return ""
        r = round(f)
        if abs(f - r) < 1e-9 and 0 <= r < 10**15:
            return str(int(r))
    s = str(val).strip()
    return re.sub(r"\D", "", s)


def acumulado_por_contraparte(
    df_ajustado: pd.DataFrame, emitidos: bool
) -> list[dict[str, Any]]:
    """
    Un acumulado por proveedor (recibidos) o cliente (emitidos): Nombre, CUIT,
    Neto (cuatro conceptos de neto), IVA (Total IVA) y total (Imp. Total).
    Valores de ``df_ajustado`` (signo y tipo de cambio ya aplicados).
    """
    col_nom = "Denominación Receptor" if emitidos else "Denominación Emisor"
    col_doc = "Nro. Doc. Receptor" if emitidos else "Nro. Doc. Emisor"
    for c in [col_nom, col_doc, "Total IVA", "Imp. Total"] + COLUMNAS_NETO_INFORME_CONTRAPARTE:
        if c not in df_ajustado.columns:
            return []

    d = df_ajustado[
        [col_nom, col_doc, "Total IVA", "Imp. Total"] + COLUMNAS_NETO_INFORME_CONTRAPARTE
    ].copy()
    d["_k"] = d[col_doc].map(_normalizar_clave_cuit_doc)
    neto_sum = (
        d[COLUMNAS_NETO_INFORME_CONTRAPARTE]
        .apply(pd.to_numeric, errors="coerce")
        .fillna(0)
        .sum(axis=1)
    )
    d["_neto"] = neto_sum
    d["_iva"] = pd.to_numeric(d["Total IVA"], errors="coerce").fillna(0)
    d["_tot"] = pd.to_numeric(d["Imp. Total"], errors="coerce").fillna(0)

    out: list[dict[str, Any]] = []
    for cuit_key, grp in d.groupby("_k", sort=False):
        noms = grp[col_nom].astype(str).str.strip()
        noms = noms.replace("", pd.NA).replace("nan", pd.NA)
        if noms.notna().any():
            mod = noms.mode()
            nombre = str(mod.iloc[0]) if len(mod) else str(noms.dropna().iloc[0])
        else:
            nombre = ""
        # En pantalla y Excel: solo dígitos (8 DNI, 11 CUIT), sin puntos ni comas
        cuit_solo = cuit_key if cuit_key else ""
        out.append(
            {
                "nombre": nombre,
                "cuit": cuit_solo,
                "neto": float(grp["_neto"].sum()),
                "iva": float(grp["_iva"].sum()),
                "total": float(grp["_tot"].sum()),
            }
        )
    out.sort(
        key=lambda r: ((r["nombre"] or "zzz").lower(), r["cuit"] or "")
    )
    return out


# DataFrame de imputaciones: columnas (cuit, razón|None, código, nombre) ya resueltas
ATTR_IMPUTACION_COLUMNAS = "imputacion_resuelta_cols"


def _es_celda_candidata_documento_cuit(val: object) -> bool:
    """True si el valor puede ser un CUIT/DNI normalizado (solo dígitos, longitud habitual)."""
    key = _normalizar_clave_cuit_doc(val)
    if not key or not key.isdigit():
        return False
    n = len(key)
    return n == 11 or (8 <= n <= 10)


def _primera_fila_y_columna_cuit_en_matriz(raw: pd.DataFrame) -> tuple[int, int] | None:
    """Primera celda (fila, columna) que parece contener un CUIT/documento. None si no hay."""
    if raw is None or raw.empty:
        return None
    nrows, ncols = int(raw.shape[0]), int(raw.shape[1])
    lim = min(nrows, 8000)
    for ri in range(lim):
        for ci in range(ncols):
            if _es_celda_candidata_documento_cuit(raw.iat[ri, ci]):
                return (ri, ci)
    return None


def _nombres_columna_unicos_desde_fila_encabezado(hdr_vals: list) -> list[str]:
    """Evita nombres duplicados (p. ej. varias celdas vacías en la fila de encabezado)."""
    counts: dict[str, int] = defaultdict(int)
    out: list[str] = []
    for j, v in enumerate(hdr_vals):
        if v is None or pd.isna(v):
            base = f"col_{j}"
        else:
            base = str(v).strip()
        if not base or base.lower() == "nan":
            base = f"col_{j}"
        c = counts[base]
        counts[base] += 1
        out.append(base if c == 0 else f"{base}_{c}")
    return out


def _intentar_df_imputacion_con_encabezado_en_fila(
    raw: pd.DataFrame,
    fila_encabezado: int,
    fila_datos_desde: int,
    col_cuit_primera: int,
) -> pd.DataFrame | None:
    """
    Usa ``fila_encabezado`` como nombres de columnas y ``fila_datos_desde`` como inicio de datos.
    Si la celda (fila_encabezado, col_cuit_primera) parece un documento numérico, no se usa como encabezado.
    """
    if fila_encabezado < 0 or fila_datos_desde <= fila_encabezado:
        return None
    if fila_datos_desde >= len(raw):
        return None
    if _es_celda_candidata_documento_cuit(raw.iat[fila_encabezado, col_cuit_primera]):
        return None

    ncols = int(raw.shape[1])
    hdr_vals = [raw.iat[fila_encabezado, j] for j in range(ncols)]
    cols = _nombres_columna_unicos_desde_fila_encabezado(hdr_vals)
    body = raw.iloc[fila_datos_desde:].copy()
    if body.shape[1] != len(cols):
        return None
    body.columns = cols
    try:
        t = detectar_columnas_archivo_imputaciones(list(body.columns))
    except ValueError:
        return None
    body.attrs[ATTR_IMPUTACION_COLUMNAS] = t
    return body


def _columnas_imputacion_por_posicion_rel_cuit(
    col_cuit: int, ncols: int
) -> tuple[str, str | None, str, str]:
    """
    Sin encabezados reconocibles: ``col_cuit`` = índice de columna CUIT.
    A la derecha: 2 columnas (código, nombre) o 3+ (opcional razón, código, nombre).
    """
    if col_cuit < 0 or col_cuit >= ncols:
        raise ValueError("imput_pos")
    k = ncols - col_cuit - 1
    if k < 2:
        raise ValueError("imput_cols")
    cuit_s = str(col_cuit)
    if k == 2:
        return (cuit_s, None, str(col_cuit + 1), str(col_cuit + 2))
    return (cuit_s, str(col_cuit + 1), str(col_cuit + 2), str(col_cuit + 3))


def _df_imputacion_desde_matriz_posicional(
    raw: pd.DataFrame, fila_datos_desde: int, col_cuit: int
) -> pd.DataFrame:
    ncols = int(raw.shape[1])
    body = raw.iloc[fila_datos_desde:].copy()
    body.columns = [str(i) for i in range(ncols)]
    t = _columnas_imputacion_por_posicion_rel_cuit(col_cuit, ncols)
    body.attrs[ATTR_IMPUTACION_COLUMNAS] = t
    return body


def _dataframe_imputacion_desde_matriz_bruta(
    raw: pd.DataFrame, ui_lang: str
) -> pd.DataFrame | None:
    """
    Localiza la primera fila con un CUIT/documento y arma el DataFrame (con o sin fila de títulos).
    Devuelve None si no hay ningún candidato a CUIT.
    """
    pos = _primera_fila_y_columna_cuit_en_matriz(raw)
    if pos is None:
        return None
    r, ic = pos

    if r > 0:
        df_hdr = _intentar_df_imputacion_con_encabezado_en_fila(raw, r - 1, r, ic)
        if df_hdr is not None:
            return df_hdr

    try:
        return _df_imputacion_desde_matriz_posicional(raw, r, ic)
    except ValueError as exc:
        if str(exc) == "imput_cols":
            raise ValueError(
                _mensaje_procesamiento(
                    ui_lang,
                    es=(
                        "Archivo de imputaciones: a la derecha del primer CUIT detectado "
                        "hacen falta al menos 2 columnas (código y nombre de cuenta)."
                    ),
                    en=(
                        "Imputations file: at least 2 columns to the right of the first "
                        "detected CUIT are required (code and account name)."
                    ),
                )
            ) from exc
        if str(exc) == "imput_pos":
            raise ValueError(
                _mensaje_procesamiento(
                    ui_lang,
                    es="Archivo de imputaciones: posición de CUIT inválida.",
                    en="Imputations file: invalid CUIT column position.",
                )
            ) from exc
        raise


def _header_sin_acentos(h: object) -> str:
    s = str(h).strip()
    s = "".join(
        c
        for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )
    return s.lower()


def _score_rol_columna_imputaciones(h: object) -> dict[str, float]:
    """Puntuación por rol de columna para detectar encabezados del archivo de imputación."""
    x = _header_sin_acentos(h)
    return {
        "cuit": (
            (25.0 if "cuit" in x else 0.0)
            + (18.0 if "nro" in x and "doc" in x else 0.0)
            + (12.0 if "documento" in x and "tipo" not in x else 0.0)
        ),
        "razon": (
            (22.0 if "razon" in x and "social" in x else 0.0)
            + (16.0 if "denominacion" in x and "imput" not in x else 0.0)
            + (10.0 if x in ("nombre", "nombre fantasia", "fantasia") else 0.0)
            + (8.0 if "proveedor" in x or "cliente" in x else 0.0)
        ),
        "codigo": (
            (28.0 if "imput" in x and ("cod" in x or "cuenta" in x or "cta" in x) else 0.0)
            + (20.0 if "cod" in x and "contab" in x else 0.0)
            + (14.0 if x.startswith("cod") and "imput" in x else 0.0)
            + (10.0 if "plan" in x and "cuenta" in x else 0.0)
            + (12.0 if x in ("cod", "codigo", "cód", "código", "cta", "nro cuenta", "nº cuenta") else 0.0)
            + (8.0 if x.startswith("cod") and len(x) <= 6 else 0.0)
        ),
        "nombre": (
            (26.0 if "imput" in x and ("nombre" in x or "descrip" in x or "denominacion" in x) else 0.0)
            + (18.0 if "nombre" in x and "cuenta" in x else 0.0)
            + (14.0 if "descrip" in x and "cuenta" in x else 0.0)
            + (10.0 if x in ("nombre", "descripcion", "descripción", "nom", "detalle") else 0.0)
        ),
    }


def _elegir_columna(
    columnas: list[str], rol: str, prohibidas: set[str]
) -> tuple[str | None, float]:
    mejor_c: str | None = None
    mejor_s = -1.0
    for c in columnas:
        if c in prohibidas:
            continue
        s = _score_rol_columna_imputaciones(c)[rol]
        if s > mejor_s:
            mejor_s = s
            mejor_c = c
    return mejor_c, mejor_s


def detectar_columnas_archivo_imputaciones(
    columnas: list[str],
) -> tuple[str, str | None, str, str]:
    """
    Devuelve (col_cuit, col_razon_social|None, col_codigo, col_nombre_cuenta).
    ``col_razon_social`` es opcional; si no se detecta con suficiente confianza, None.
    """
    cols = [str(c).strip() for c in columnas if str(c).strip()]
    if len(cols) < 3:
        raise ValueError("imput_cols")

    prohibidas: set[str] = set()
    cuit_c, sc = _elegir_columna(cols, "cuit", prohibidas)
    if cuit_c is None or sc < 8.0:
        raise ValueError("imput_cuit")
    prohibidas.add(cuit_c)

    cod_c, sco = _elegir_columna(cols, "codigo", prohibidas)
    if cod_c is None or sco < 8.0:
        raise ValueError("imput_codigo")
    prohibidas.add(cod_c)

    nom_c, sn = _elegir_columna(cols, "nombre", prohibidas)
    if nom_c is None or sn < 6.0:
        raise ValueError("imput_nombre")
    prohibidas.add(nom_c)

    raz_c, sr = _elegir_columna(cols, "razon", prohibidas)
    razon_final = raz_c if raz_c is not None and sr >= 8.0 else None

    return cuit_c, razon_final, cod_c, nom_c


def _leer_excel_matriz_sin_encabezado(entrada) -> pd.DataFrame | None:
    if hasattr(entrada, "seek"):
        entrada.seek(0)
    try:
        raw = pd.read_excel(entrada, sheet_name=0, header=None)
    except Exception:
        return None
    if raw is None or raw.empty:
        return None
    return raw


def _leer_excel_tabla_imputaciones_encabezado_clasico(entrada, ui_lang: str) -> pd.DataFrame:
    """Compatibilidad: primera o segunda fila del .xlsx como encabezado con nombres reconocibles."""
    mejor: pd.DataFrame | None = None
    for header_row in (0, 1):
        if hasattr(entrada, "seek"):
            entrada.seek(0)
        try:
            raw = pd.read_excel(entrada, sheet_name=0, header=header_row)
        except Exception:
            continue
        raw.columns = raw.columns.astype(str).str.strip()
        if raw.shape[1] < 2:
            continue
        ok = False
        try:
            detectar_columnas_archivo_imputaciones(list(raw.columns))
            ok = True
        except ValueError:
            pass
        if ok:
            return raw
        if mejor is None or raw.shape[1] > mejor.shape[1]:
            mejor = raw

    if mejor is not None:
        return mejor
    raise ValueError(
        _mensaje_procesamiento(
            ui_lang,
            es="No se pudo leer el Excel de imputaciones.",
            en="Could not read the imputations Excel file.",
        )
    )


def _leer_excel_tabla_imputaciones(entrada, ui_lang: str) -> pd.DataFrame:
    if hasattr(entrada, "seek"):
        entrada.seek(0)
    mat = _leer_excel_matriz_sin_encabezado(entrada)
    if mat is not None:
        try:
            df = _dataframe_imputacion_desde_matriz_bruta(mat, ui_lang)
            if df is not None:
                return df
        except ValueError:
            raise
        except Exception:
            pass
    if hasattr(entrada, "seek"):
        entrada.seek(0)
    return _leer_excel_tabla_imputaciones_encabezado_clasico(entrada, ui_lang)


def _leer_csv_tabla_imputaciones(
    entrada,
    nombre_archivo: str | None,
    ui_lang: str,
) -> pd.DataFrame:
    muestra = ""
    try:
        if hasattr(entrada, "seek"):
            entrada.seek(0)
        if hasattr(entrada, "read"):
            raw_b = entrada.read(8192)
            muestra = (
                raw_b.decode("utf-8", errors="ignore")
                if isinstance(raw_b, bytes)
                else str(raw_b)
            )
    finally:
        if hasattr(entrada, "seek"):
            entrada.seek(0)

    delimitadores = [";", ",", "\t", "|"]
    skiprows_opciones = [0]
    lineas = [ln.strip() for ln in muestra.splitlines() if ln.strip()]
    primera = lineas[0] if lineas else ""
    if primera.lower().startswith("sep="):
        sep_decl = primera.split("=", 1)[1].strip()
        if sep_decl:
            delim_decl = sep_decl[0]
            delimitadores = [delim_decl] + [d for d in delimitadores if d != delim_decl]
        skiprows_opciones = [1, 0]
    else:
        try:
            dialecto = csv.Sniffer().sniff(muestra or "", delimiters=";,|\t,")
            delim_sniff = dialecto.delimiter
            delimitadores = [delim_sniff] + [d for d in delimitadores if d != delim_sniff]
        except Exception:
            pass

    def _intentar_matriz_bruta() -> pd.DataFrame | None:
        for skiprows in skiprows_opciones:
            for delimitador in delimitadores:
                for on_bad_lines in (None, "skip"):
                    if hasattr(entrada, "seek"):
                        entrada.seek(0)
                    kwargs: dict[str, Any] = {
                        "header": None,
                        "skiprows": skiprows,
                        "sep": delimitador,
                        "engine": "python",
                        "skipinitialspace": True,
                        "encoding": "utf-8-sig",
                        "dayfirst": True,
                    }
                    if on_bad_lines is not None:
                        kwargs["on_bad_lines"] = on_bad_lines
                    try:
                        mat = pd.read_csv(entrada, **kwargs)
                    except (pd.errors.ParserError, Exception):
                        continue
                    if mat.shape[1] < 2:
                        continue
                    try:
                        return _dataframe_imputacion_desde_matriz_bruta(mat, ui_lang)
                    except ValueError:
                        raise
                    except Exception:
                        continue
        return None

    try:
        df_mat = _intentar_matriz_bruta()
        if df_mat is not None:
            return df_mat
    except ValueError:
        raise

    primer_df = None
    for skiprows in skiprows_opciones:
        for delimitador in delimitadores:
            for on_bad_lines in (None, "skip"):
                if hasattr(entrada, "seek"):
                    entrada.seek(0)
                kwargs: dict[str, Any] = {
                    "header": 0,
                    "skiprows": skiprows,
                    "sep": delimitador,
                    "engine": "python",
                    "skipinitialspace": True,
                    "encoding": "utf-8-sig",
                    "dayfirst": True,
                }
                if on_bad_lines is not None:
                    kwargs["on_bad_lines"] = on_bad_lines
                try:
                    candidato = pd.read_csv(entrada, **kwargs)
                except (pd.errors.ParserError, Exception):
                    continue
                candidato.columns = candidato.columns.astype(str).str.strip()
                if candidato.shape[1] < 2:
                    continue
                if primer_df is None:
                    primer_df = candidato
                try:
                    detectar_columnas_archivo_imputaciones(list(candidato.columns))
                    return candidato
                except ValueError:
                    continue

    if primer_df is not None:
        return primer_df
    raise ValueError(
        _mensaje_procesamiento(
            ui_lang,
            es="No se pudo leer el CSV de imputaciones.",
            en="Could not read the imputations CSV file.",
        )
    )


def leer_dataframe_imputaciones(
    entrada: str | io.BytesIO,
    nombre_archivo: str | None = None,
    ui_lang: str = "en",
) -> pd.DataFrame:
    """Lee .xlsx o .csv de referencia (CUIT + imputación) sin normalizar columnas de comprobantes."""
    nombre = (nombre_archivo or str(entrada)).lower()
    if nombre.endswith(".csv"):
        return _leer_csv_tabla_imputaciones(entrada, nombre_archivo, ui_lang)
    return _leer_excel_tabla_imputaciones(entrada, ui_lang)


def construir_mapa_cuit_a_imputacion(
    df: pd.DataFrame,
    *,
    cuit_col: str,
    codigo_col: str,
    nombre_col: str,
) -> dict[str, tuple[str, str]]:
    """
    CUIT normalizado (solo dígitos) -> (código imputación, nombre cuenta).
    Si un CUIT aparece varias veces, gana la primera fila del archivo (orden de lectura).
    """

    def _texto_celda(val: object) -> str:
        if val is None or pd.isna(val):
            return ""
        if isinstance(val, (float, np.floating)):
            f = float(val)
            if np.isfinite(f) and abs(f - round(f)) < 1e-9:
                return str(int(round(f)))
        if isinstance(val, (int, np.integer)):
            return str(int(val))
        return str(val).strip()

    out: dict[str, tuple[str, str]] = {}
    for _, row in df.iterrows():
        key = _normalizar_clave_cuit_doc(row.get(cuit_col))
        if not key:
            continue
        if key in out:
            continue
        cod_s = _texto_celda(row.get(codigo_col))
        nom_s = _texto_celda(row.get(nombre_col))
        out[key] = (cod_s, nom_s)
    return out


def leer_mapa_imputaciones_desde_archivo(
    entrada: str | io.BytesIO,
    nombre_archivo: str | None = None,
    ui_lang: str = "en",
) -> dict[str, tuple[str, str]]:
    """
    Lee el archivo opcional de referencia y devuelve el mapa CUIT -> (código, nombre imputación).
    """
    df = leer_dataframe_imputaciones(entrada, nombre_archivo=nombre_archivo, ui_lang=ui_lang)
    resuelto = df.attrs.get(ATTR_IMPUTACION_COLUMNAS)
    if resuelto is not None:
        cuit_c, _raz, cod_c, nom_c = resuelto
    else:
        try:
            cuit_c, _raz, cod_c, nom_c = detectar_columnas_archivo_imputaciones(
                list(df.columns)
            )
        except ValueError as exc:
            if str(exc) == "imput_cols":
                msg_es = (
                    "Archivo de imputaciones: se necesitan al menos 3 columnas "
                    "(CUIT, código y nombre de imputación)."
                )
                msg_en = (
                    "Imputations file: at least 3 columns are required "
                    "(tax ID, imputation code and account name)."
                )
            elif str(exc) == "imput_cuit":
                msg_es = (
                    "Archivo de imputaciones: no se detectó una columna de CUIT/documento. "
                    "Usá un encabezado que contenga «CUIT» o «Nro. Doc.», o una primera fila de datos con el CUIT."
                )
                msg_en = (
                    "Imputations file: no CUIT/document column was detected. "
                    "Use a header containing «CUIT» or similar, or a first data row with the tax ID."
                )
            elif str(exc) == "imput_codigo":
                msg_es = (
                    "Archivo de imputaciones: no se detectó la columna de código de imputación contable."
                )
                msg_en = (
                    "Imputations file: accounting imputation code column was not detected."
                )
            else:
                msg_es = (
                    "Archivo de imputaciones: no se detectó la columna de nombre de la cuenta de imputación."
                )
                msg_en = (
                    "Imputations file: imputation account name column was not detected."
                )
            raise ValueError(_mensaje_procesamiento(ui_lang, es=msg_es, en=msg_en)) from exc

    return construir_mapa_cuit_a_imputacion(
        df, cuit_col=cuit_c, codigo_col=cod_c, nombre_col=nom_c
    )


def enriquecer_contrapartes_con_imputacion(
    tabla_contrapartes: list[dict[str, Any]],
    mapa_cuit_imputacion: dict[str, tuple[str, str]] | None,
) -> list[dict[str, Any]]:
    """Añade ``codigo_imputacion`` y ``nombre_imputacion`` a cada fila de contraparte."""
    if not mapa_cuit_imputacion:
        out = []
        for r in tabla_contrapartes:
            row = dict(r)
            row.setdefault("codigo_imputacion", "")
            row.setdefault("nombre_imputacion", "")
            out.append(row)
        return out

    out: list[dict[str, Any]] = []
    for r in tabla_contrapartes:
        row = dict(r)
        k = _normalizar_clave_cuit_doc(row.get("cuit", ""))
        cod, nom = mapa_cuit_imputacion.get(k, ("", ""))
        row["codigo_imputacion"] = cod
        row["nombre_imputacion"] = nom
        out.append(row)
    return out


def resumen_totales_por_imputacion(
    tabla_contrapartes_enriquecida: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """
    Agrupa por código y nombre de imputación; suma neto, IVA y total.
    Las filas sin código ni nombre quedan en un mismo grupo (vacío).
    """
    acc: dict[tuple[str, str], dict[str, Any]] = {}
    for r in tabla_contrapartes_enriquecida:
        cod = str(r.get("codigo_imputacion", "") or "").strip()
        nom = str(r.get("nombre_imputacion", "") or "").strip()
        key = (cod, nom)
        if key not in acc:
            acc[key] = {
                "codigo_imputacion": cod,
                "nombre_imputacion": nom,
                "neto": 0.0,
                "iva": 0.0,
                "total": 0.0,
            }
        b = acc[key]
        b["neto"] += float(r.get("neto", 0) or 0)
        b["iva"] += float(r.get("iva", 0) or 0)
        b["total"] += float(r.get("total", 0) or 0)

    filas = list(acc.values())
    filas.sort(
        key=lambda x: (
            (x["codigo_imputacion"] or "\uffff").lower(),
            (x["nombre_imputacion"] or "").lower(),
        )
    )
    return filas


def _mensaje_procesamiento(ui_lang: str, *, es: str, en: str) -> str:
    """Errores de validación: español solo si la UI es es; en cualquier otro idioma, inglés."""
    return es if ui_lang == "es" else en


def _mejor_dataframe_excel(entrada, hoja: str | int, ui_lang: str = "en") -> pd.DataFrame:
    """
    Prueba header=0 y header=1 en .xlsx y elige el que tenga todas las columnas requeridas.
    Si ninguno las tiene todas, elige el que menos falte (misma lógica que CSV vs xlsx).
    """
    req = _columnas_requeridas_lectura()
    opciones: list[tuple[int, int, pd.DataFrame]] = []

    for header_row in (0, 1):
        if hasattr(entrada, "seek"):
            entrada.seek(0)
        try:
            raw = pd.read_excel(entrada, sheet_name=hoja, header=header_row)
        except Exception:
            continue
        raw.columns = raw.columns.astype(str).str.strip()
        cand = normalizar_columnas(raw)
        faltan = len(req - set(cand.columns))
        opciones.append((faltan, header_row, cand))

    if not opciones:
        raise ValueError(
            _mensaje_procesamiento(
                ui_lang,
                es="No se pudo leer el archivo Excel.",
                en="Could not read the Excel file.",
            )
        )

    opciones.sort(key=lambda t: (t[0], t[1]))
    mejor_faltan, _hdr, mejor_df = opciones[0]
    if mejor_faltan > 0:
        nombres = ", ".join(mejor_df.columns.astype(str))
        faltantes = list(req - set(mejor_df.columns))
        raise ValueError(
            _mensaje_procesamiento(
                ui_lang,
                es=(
                    f"No se encontraron las columnas: {faltantes}. "
                    f"Columnas en el archivo: {nombres}"
                ),
                en=(
                    f"Missing columns: {faltantes}. "
                    f"Columns in the file: {nombres}"
                ),
            )
        )
    return mejor_df


def leer_tabla(
    entrada,
    hoja: str | int = 0,
    nombre_archivo: str | None = None,
    ui_lang: str = "en",
) -> pd.DataFrame:
    """
    Lee un .xlsx o .csv con formato:
    - .xlsx: fila 1 encabezado general, fila 2 encabezados de columnas, fila 3+ datos
    - .csv: fila 1 encabezados de columnas, fila 2+ datos
    """
    nombre = (nombre_archivo or str(entrada)).lower()
    if nombre.endswith(".csv"):
        # CSV: encabezados en fila 1 (excepto archivos con primera línea "sep=;")
        # Se prueban varias combinaciones y se elige la que contiene columnas requeridas.
        columnas_requeridas = _columnas_requeridas_lectura()
        muestra = ""
        delimitadores = [";", ",", "\t", "|"]
        skiprows_opciones = [0]

        try:
            if hasattr(entrada, "seek"):
                entrada.seek(0)
            if hasattr(entrada, "read"):
                muestra = entrada.read(8192)
                if isinstance(muestra, bytes):
                    muestra = muestra.decode("utf-8", errors="ignore")
        finally:
            if hasattr(entrada, "seek"):
                entrada.seek(0)

        lineas = [ln.strip() for ln in muestra.splitlines() if ln.strip()]
        primera = lineas[0] if lineas else ""
        if primera.lower().startswith("sep="):
            sep_decl = primera.split("=", 1)[1].strip()
            if sep_decl:
                delim_decl = sep_decl[0]
                delimitadores = [delim_decl] + [d for d in delimitadores if d != delim_decl]
            skiprows_opciones = [1, 0]
        else:
            try:
                dialecto = csv.Sniffer().sniff(muestra or "", delimiters=";,|\t,")
                delim_sniff = dialecto.delimiter
                delimitadores = [delim_sniff] + [d for d in delimitadores if d != delim_sniff]
            except Exception:
                pass

        primer_df = None
        df = None
        for skiprows in skiprows_opciones:
            for delimitador in delimitadores:
                for on_bad_lines in (None, "skip"):
                    if hasattr(entrada, "seek"):
                        entrada.seek(0)
                    kwargs = {
                        "header": 0,
                        "skiprows": skiprows,
                        "sep": delimitador,
                        "engine": "python",
                        "skipinitialspace": True,
                        "encoding": "utf-8-sig",
                        # ARCA / Argentina: día primero en fechas ambiguas al inferir dtypes
                        "dayfirst": True,
                    }
                    if on_bad_lines is not None:
                        kwargs["on_bad_lines"] = on_bad_lines
                    try:
                        candidato = pd.read_csv(entrada, **kwargs)
                    except pd.errors.ParserError:
                        continue
                    except Exception:
                        continue

                    candidato.columns = candidato.columns.astype(str).str.strip()
                    candidato = normalizar_columnas(candidato)
                    if primer_df is None:
                        primer_df = candidato
                    if columnas_requeridas.issubset(set(candidato.columns)):
                        df = candidato
                        break
                if df is not None:
                    break
            if df is not None:
                break

        if df is None:
            # Fallback: primera lectura exitosa aunque no tenga todas las columnas.
            if primer_df is not None:
                df = primer_df
            else:
                raise ValueError(
                    _mensaje_procesamiento(
                        ui_lang,
                        es="No se pudo leer el CSV con un formato válido.",
                        en="Could not read the CSV in a valid format.",
                    )
                )
    else:
        df = _mejor_dataframe_excel(entrada, hoja, ui_lang=ui_lang)

    df.columns = df.columns.astype(str).str.strip()
    return normalizar_columnas(df)


def procesar_archivo(
    ruta_excel: str | io.BytesIO,
    hoja: str | int = 0,
    nombre_archivo: str | None = None,
    ui_lang: str = "en",
    *,
    emitidos: bool = False,
) -> tuple[
    pd.DataFrame,
    dict[str, float],
    dict[str, dict[str, float]],
    dict,
    list[dict[str, Any]],
]:
    """
    Lee un archivo Excel y devuelve la sumatoria de las columnas indicadas.
    Fila 1 = encabezado general, fila 2 = encabezados de columnas, datos desde fila 3.
    Las filas con Tipo = nota de crédito se suman en valor negativo.
    Comprobantes recibidos (emitidos=False): tipos en CODIGOS_IMP_TOTAL_EN_NETO_IVA_0 (B/C y afines):
    el importe suele estar solo en Imp. Total; se refleja en Neto Grav. IVA 0% y en Imp. Total,
    Neto Gravado Total en 0 y el resto de columnas numéricas de esa fila en 0 (antes de signo NC y TC).

    Comprobantes emitidos (emitidos=True): solo los de clase C (CODIGOS_IMP_TOTAL_EN_NETO_IVA_0_SOLO_C)
    siguen esa regla; el resto suma con los valores de cada columna del archivo (Factura A/B, FCE A/B, etc.).

    Args:
        ruta_excel: Ruta al archivo .xlsx o buffer legible por pandas (p. ej. BytesIO)
        hoja: Nombre o índice de la hoja (0 por defecto)
        emitidos: True si el archivo es export de comprobantes emitidos.

    Returns:
        Tuple con:
        - DataFrame ajustado (columnas numéricas con signo aplicado según Tipo)
        - Diccionario con el nombre de cada columna y su suma.
        - Diccionario período (``\"YYYY-MM\"``) -> totales por columna.
        - Diccionario con «Neto/IVA notas de crédito» (anual y por período) alícuotas 2,5%–27%.
        - Tabla de acumulados por proveedor o cliente (Nombre, CUIT, Neto, IVA, Total).
    """
    # header=1: la fila 2 del archivo (índice 1) tiene los nombres de columnas; datos desde fila 3
    df = leer_tabla(
        ruta_excel, hoja=hoja, nombre_archivo=nombre_archivo, ui_lang=ui_lang
    )

    # Comprobar que existan todas las columnas necesarias
    columnas_requeridas = COLUMNAS_A_AJUSTAR + ["Tipo", "Tipo Cambio", "Fecha Emisión"]
    faltantes = [c for c in columnas_requeridas if c not in df.columns]
    if faltantes:
        nombres = ", ".join(df.columns.astype(str))
        raise ValueError(
            _mensaje_procesamiento(
                ui_lang,
                es=(
                    f"No se encontraron las columnas: {faltantes}. "
                    f"Columnas en el archivo: {nombres}"
                ),
                en=(
                    f"Missing columns: {faltantes}. "
                    f"Columns in the file: {nombres}"
                ),
            )
        )

    df = df.reset_index(drop=True)

    # Signo: -1 si es nota de crédito, +1 si no (código numérico AFIP)
    codigo_num = serie_codigo_tipo_comprobante(df["Tipo"])
    es_nota_credito = codigo_num.isin(CODIGOS_NOTA_CREDITO)
    signo = (1 - 2 * es_nota_credito.astype(int)).reset_index(drop=True)

    codigos_neto_iva_0 = (
        CODIGOS_IMP_TOTAL_EN_NETO_IVA_0_SOLO_C
        if emitidos
        else CODIGOS_IMP_TOTAL_EN_NETO_IVA_0
    )
    es_imp_en_neto_iva_0 = codigo_num.isin(codigos_neto_iva_0).reset_index(drop=True)

    # Factor de conversión por fila: vacíos/no numéricos se toman como 0 solo para cálculo
    tipo_cambio = serie_a_float_importe(df["Tipo Cambio"]).fillna(0).reset_index(
        drop=True
    )

    # Ajustar signos y tipo de cambio en el DataFrame de salida, luego acumular totales
    df_ajustado = df.copy()
    resultado: dict[str, float] = {}
    imp_total_num = serie_a_float_importe(df["Imp. Total"]).fillna(0).reset_index(
        drop=True
    )
    neto_grav_num = serie_a_float_importe(df["Neto Gravado Total"]).fillna(0).reset_index(
        drop=True
    )
    neto_iva0_num = serie_a_float_importe(df["Neto Grav. IVA 0%"]).fillna(0).reset_index(
        drop=True
    )

    for col in COLUMNAS_A_AJUSTAR:
        if col == "Neto Grav. IVA 0%":
            valores = neto_iva0_num.where(~es_imp_en_neto_iva_0, imp_total_num)
        elif col == "Neto Gravado Total":
            valores = neto_grav_num.where(~es_imp_en_neto_iva_0, 0.0)
        elif col == "Imp. Total":
            # Misma base que Neto Grav. IVA 0% en B/C: coherencia tras signo y tipo de cambio
            valores = imp_total_num
        else:
            base = serie_a_float_importe(df[col]).fillna(0).reset_index(drop=True)
            valores = base.where(~es_imp_en_neto_iva_0, 0.0)
        valores_ajustados = (valores * signo * tipo_cambio).astype(float)
        df_ajustado[col] = valores_ajustados.values

    resultado, totales_por_periodo = _totales_anuales_y_por_periodo(
        df_ajustado, COLUMNAS_A_AJUSTAR, nombre_archivo
    )
    t_neto_nc, t_iva_nc, neto_nc_p, iva_nc_p = totales_notas_credito_neto_e_iva_por_periodo(
        df_ajustado, nombre_archivo
    )
    notas_credito_extras = {
        "total_neto_nc": t_neto_nc,
        "total_iva_nc": t_iva_nc,
        "neto_nc_por_periodo": neto_nc_p,
        "iva_nc_por_periodo": iva_nc_p,
    }
    tabla_contrapartes = acumulado_por_contraparte(df_ajustado, emitidos)
    _formatear_fecha_emision_salida_excel(df_ajustado)

    return (
        df_ajustado,
        resultado,
        totales_por_periodo,
        notas_credito_extras,
        tabla_contrapartes,
    )


# Contabilidad (Excel): alineación, negativos entre paréntesis, cero como guión; sin símbolo de moneda.
# Equivalente a “Contabilidad” sin divisa en la cinta de Excel.
_FORMATO_CONTABILIDAD_SIN_MONEDA = (
    r'_ * #,##0.00_ ;_ * (#,##0.00)_ ;_ * "-"??_ ;_ @_ '
)

_EPS_CERO = 1e-12


def _es_cero_numerico(val) -> bool:
    if val is None or val == "":
        return False
    if type(val) is bool:
        return False
    if isinstance(val, (int, float)):
        return abs(float(val)) < _EPS_CERO
    return False


def _longitud_texto_celda_excel(val) -> int:
    """Longitud aproximada del texto mostrado (encabezados; importes como en contabilidad sin moneda)."""
    if val is None or val == "":
        return 0
    if type(val) is bool:
        return len(str(val))
    if isinstance(val, (int, float)):
        if _es_cero_numerico(val):
            return 3
        fv = float(val)
        if fv < 0:
            return len(f"({abs(fv):,.2f})")
        return len(f"{fv:,.2f}")
    return len(str(val))


# Columnas a ocultar en hoja de comprobantes (exporte)
_COLUMNA_OCULTA_COMUN = [
    "Número Hasta",
    "Cód. Autorización",
    "Tipo Doc. Emisor",
    "Tipo Doc. Receptor",
]
_COLUMNA_OCULTA_SI_RECIBIDOS = ["Nro. Doc. Receptor"]
_ANCHO_DENOM = 40
_SHEET_COMPR = "Comprobantes"


def _nombres_columnas_ocultar(emitidos: bool) -> set[str]:
    s = set(_COLUMNA_OCULTA_COMUN)
    if not emitidos:
        s |= set(_COLUMNA_OCULTA_SI_RECIBIDOS)
    return s


def _aplicar_hoja_comprobantes_excel(
    _wb,
    ws,
    encabezados: list,
    emitidos: bool,
    titulo_hoja: str | None = None,
) -> None:
    negrita = Font(bold=True)
    for cell in ws[1]:
        cell.font = negrita

    ocultar = _nombres_columnas_ocultar(emitidos)
    denom_anchura = {
        "Denominación Emisor",
        "Denominación Receptor",
    }

    for col_i, nombre in enumerate(encabezados, start=1):
        if not nombre:
            continue
        letra = get_column_letter(col_i)
        dim = ws.column_dimensions[letra]
        if nombre in ocultar:
            dim.hidden = True
            dim.width = 0.5
            continue
        if nombre in denom_anchura:
            dim.width = float(_ANCHO_DENOM)
            for fila in range(2, ws.max_row + 1):
                c = ws.cell(row=fila, column=col_i)
                if c.value and isinstance(c.value, str) and len(c.value) > _ANCHO_DENOM * 1.2:
                    c.alignment = Alignment(wrap_text=True, vertical="top")
            continue

        max_long = 0
        for fila in range(1, ws.max_row + 1):
            v = ws.cell(row=fila, column=col_i).value
            max_long = max(max_long, _longitud_texto_celda_excel(v), len(str(nombre) if nombre else ""))
        if max_long > 0:
            dim.width = min(max_long + 2, 60)

    for col_i, nombre in enumerate(encabezados, start=1):
        if nombre in COLUMNAS_A_AJUSTAR:
            for fila in range(2, ws.max_row + 1):
                cel = ws.cell(row=fila, column=col_i)
                if cel.value is not None and cel.value != "":
                    cel.number_format = _FORMATO_CONTABILIDAD_SIN_MONEDA

    ws.title = titulo_hoja if titulo_hoja is not None else _SHEET_COMPR
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ult = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{ult}{ws.max_row}"


def _celda_num(ws, r: int, c: int, v: object) -> None:
    cl = ws.cell(row=r, column=c, value=v)
    if v is not None and isinstance(v, (int, float)) and v != "":
        cl.number_format = _FORMATO_CONTABILIDAD_SIN_MONEDA


def _fila_monto(
    wsr,
    fila: int,
    etiq: str,
    valor: object,
    bold_l: bool = False,
) -> int:
    c1 = wsr.cell(row=fila, column=1, value=etiq)
    if bold_l:
        c1.font = Font(bold=True)
    cl = wsr.cell(row=fila, column=2, value=valor)
    if isinstance(valor, (int, float)):
        cl.number_format = _FORMATO_CONTABILIDAD_SIN_MONEDA
    return fila + 1


def _ajustar_anchos_hoja_resumen_excel(wsr) -> None:
    """Carga columnas A (texto) y B (importe) a partir del contenido real."""
    max_a = 16.0
    max_b = 12.0
    for r in range(1, wsr.max_row + 1):
        a = wsr.cell(row=r, column=1).value
        b = wsr.cell(row=r, column=2).value
        if a is not None and str(a).strip() != "":
            max_a = max(max_a, float(len(str(a))))
        if b is not None and b != "":
            max_b = max(max_b, float(_longitud_texto_celda_excel(b)))
    wsr.column_dimensions["A"].width = min(max(18.0, max_a + 1.5), 78.0)
    wsr.column_dimensions["B"].width = min(max(14.0, max_b + 2.0), 26.0)


def _ajustar_anchos_hoja_distribucion_excel(wsd) -> None:
    """Ancho mínimo para conceptos largos; columnas de períodos según importes."""
    max_a = 18.0
    for r in range(1, wsd.max_row + 1):
        v = wsd.cell(row=r, column=1).value
        if v is not None and str(v).strip() != "":
            max_a = max(max_a, float(len(str(v))))
    wsd.column_dimensions["A"].width = min(max(20.0, max_a + 1.2), 78.0)
    for c in range(2, wsd.max_column + 1):
        le = get_column_letter(c)
        maxw = 11.0
        for r in range(1, wsd.max_row + 1):
            val = wsd.cell(row=r, column=c).value
            if val is not None and str(val) != "":
                maxw = max(maxw, float(_longitud_texto_celda_excel(val)))
        wsd.column_dimensions[le].width = min(max(12.0, maxw + 2.0), 24.0)


def _ajustar_anchos_hoja_contrapartes_excel(wsp, *, con_imputacion: bool = False) -> None:
    max_n = 10.0
    max_c = 10.0
    max_ci = 10.0
    max_ni = 10.0
    for r in range(1, wsp.max_row + 1):
        n = wsp.cell(row=r, column=1).value
        cu = wsp.cell(row=r, column=2).value
        if n is not None and str(n).strip() != "":
            max_n = max(max_n, float(len(str(n))))
        if cu is not None and str(cu).strip() != "":
            max_c = max(max_c, float(len(str(cu))))
        if con_imputacion:
            ci = wsp.cell(row=r, column=3).value
            ni = wsp.cell(row=r, column=4).value
            if ci is not None and str(ci).strip() != "":
                max_ci = max(max_ci, float(len(str(ci))))
            if ni is not None and str(ni).strip() != "":
                max_ni = max(max_ni, float(len(str(ni))))
    wsp.column_dimensions["A"].width = min(max(14.0, max_n + 1.5), 80.0)
    wsp.column_dimensions["B"].width = min(max(11.0, max_c + 1.5), 20.0)
    if con_imputacion:
        wsp.column_dimensions["C"].width = min(max(12.0, max_ci + 1.5), 28.0)
        wsp.column_dimensions["D"].width = min(max(14.0, max_ni + 1.5), 56.0)
        for le in "EFG":
            wsp.column_dimensions[le].width = 16.0
    else:
        for le in "CDE":
            wsp.column_dimensions[le].width = 16.0


def _rellenar_hoja_resumen_excel(
    wsr,
    totales_resumen: dict,
    totales_detalle: dict,
    suma_total: float,
    notas_credito_extras: dict,
) -> None:
    negrita = Font(bold=True)
    t1 = wsr.cell(row=1, column=1, value="Resumen (base del total)")
    t1.font = negrita
    t1.alignment = Alignment(horizontal="center", vertical="center")
    wsr.merge_cells("A1:B1")
    fila = 2
    wsr.cell(row=fila, column=1, value="Concepto").font = negrita
    wsr.cell(row=fila, column=2, value="Importe").font = negrita
    fila_tabla0 = 2
    fila = 3
    for k, v in totales_resumen.items():
        wsr.cell(row=fila, column=1, value=k)
        cl = wsr.cell(row=fila, column=2, value=v)
        cl.number_format = _FORMATO_CONTABILIDAD_SIN_MONEDA
        fila += 1
    ctot1 = wsr.cell(row=fila, column=1, value="Total (resumen)")
    ctot1.font = negrita
    ctot2 = wsr.cell(row=fila, column=2, value=suma_total)
    ctot2.font = negrita
    ctot2.number_format = _FORMATO_CONTABILIDAD_SIN_MONEDA
    fila += 1
    fila = _fila_monto(wsr, fila, "Neto Notas de Crédito", notas_credito_extras.get("total_neto_nc", 0))
    fila = _fila_monto(wsr, fila, "IVA Notas de Crédito", notas_credito_extras.get("total_iva_nc", 0))
    u_res = fila - 1
    if totales_detalle:
        fila += 1
        wsr.cell(row=fila, column=1, value="Detalle por alícuota")
        wsr.cell(row=fila, column=1).font = negrita
        wsr.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
        fila += 1
        wsr.cell(row=fila, column=1, value="Concepto").font = negrita
        wsr.cell(row=fila, column=2, value="Importe").font = negrita
        fila += 1
        for k, v in totales_detalle.items():
            wsr.cell(row=fila, column=1, value=k)
            cl = wsr.cell(row=fila, column=2, value=v)
            cl.number_format = _FORMATO_CONTABILIDAD_SIN_MONEDA
            fila += 1
        u_res = fila - 1
    wsr.auto_filter.ref = f"A{fila_tabla0}:B{u_res}"
    _ajustar_anchos_hoja_resumen_excel(wsr)


def _rellenar_hoja_distribucion_excel(
    wsd,
    periodos_orden: list[str],
    columnas_orden: list[str],
    totales_por_periodo: dict[str, dict[str, float]],
    notas_credito_extras: dict,
) -> None:
    negrita = Font(bold=True)
    wsd.append(["Concepto"] + [""] * len(periodos_orden))
    for i, per in enumerate(periodos_orden, start=2):
        y, m = map(int, per.split("-"))
        mes_lbl = NOMBRES_MESES[m - 1].capitalize()
        chead = wsd.cell(row=1, column=i, value=f"{mes_lbl}\n{y}")
        chead.font = negrita
        chead.alignment = Alignment(
            wrap_text=True, horizontal="center", vertical="center"
        )
    c0 = wsd.cell(row=1, column=1, value="Concepto")
    c0.font = negrita
    c0.alignment = Alignment(horizontal="left", vertical="center")
    wsd.row_dimensions[1].height = 32
    fila = 2
    for col in columnas_orden:
        wsd.cell(row=fila, column=1, value=col)
        for j, per in enumerate(periodos_orden, start=2):
            vcel = totales_por_periodo.get(per, {}).get(col, 0.0)
            _celda_num(wsd, fila, j, vcel)
        fila += 1
    wsd.cell(row=fila, column=1, value="Total (resumen)").font = negrita
    for j, per in enumerate(periodos_orden, start=2):
        tr = 0.0
        for cname in COLUMNAS_TOTAL_RESUMEN:
            tr += float(totales_por_periodo.get(per, {}).get(cname, 0.0))
        cc = wsd.cell(row=fila, column=j, value=tr)
        cc.font = negrita
        cc.number_format = _FORMATO_CONTABILIDAD_SIN_MONEDA
    fila += 1
    wsd.cell(row=fila, column=1, value="Neto Notas de Crédito")
    ncp = notas_credito_extras.get("neto_nc_por_periodo", {})
    for j, per in enumerate(periodos_orden, start=2):
        _celda_num(wsd, fila, j, ncp.get(per, 0.0))
    fila += 1
    wsd.cell(row=fila, column=1, value="IVA Notas de Crédito")
    icp = notas_credito_extras.get("iva_nc_por_periodo", {})
    for j, per in enumerate(periodos_orden, start=2):
        _celda_num(wsd, fila, j, icp.get(per, 0.0))
    if wsd.max_column >= 1 and wsd.max_row >= 1:
        wsd.auto_filter.ref = f"A1:{get_column_letter(wsd.max_column)}{wsd.max_row}"
    _ajustar_anchos_hoja_distribucion_excel(wsd)


def _rellenar_hoja_contrapartes_excel(
    wsp,
    tabla_contrapartes: list[dict],
    *,
    con_imputacion: bool = False,
) -> None:
    negrita = Font(bold=True)
    if con_imputacion:
        wsp.append(
            [
                "Nombre",
                "CUIT",
                "Cód. imputación",
                "Imputación contable",
                "Neto",
                "IVA",
                "Total",
            ]
        )
    else:
        wsp.append(["Nombre", "CUIT", "Neto", "IVA", "Total"])
    for cell in wsp[1]:
        cell.font = negrita
    for r, trow in enumerate(tabla_contrapartes, start=2):
        wsp.cell(row=r, column=1, value=trow.get("nombre", ""))
        cdoc = wsp.cell(row=r, column=2, value=str(trow.get("cuit", "") or ""))
        cdoc.number_format = "@"
        if con_imputacion:
            wsp.cell(row=r, column=3, value=str(trow.get("codigo_imputacion", "") or ""))
            wsp.cell(row=r, column=4, value=str(trow.get("nombre_imputacion", "") or ""))
            _celda_num(wsp, r, 5, trow.get("neto", 0))
            _celda_num(wsp, r, 6, trow.get("iva", 0))
            _celda_num(wsp, r, 7, trow.get("total", 0))
        else:
            _celda_num(wsp, r, 3, trow.get("neto", 0))
            _celda_num(wsp, r, 4, trow.get("iva", 0))
            _celda_num(wsp, r, 5, trow.get("total", 0))
    end_r = max(1, len(tabla_contrapartes) + 1)
    ult_col = "G" if con_imputacion else "E"
    wsp.auto_filter.ref = f"A1:{ult_col}{end_r}"
    _ajustar_anchos_hoja_contrapartes_excel(wsp, con_imputacion=con_imputacion)


def _ajustar_anchos_hoja_resumen_imputacion_excel(wsi) -> None:
    wsi.column_dimensions["A"].width = 18.0
    wsi.column_dimensions["B"].width = 44.0
    for le in "CDE":
        wsi.column_dimensions[le].width = 16.0


def _rellenar_hoja_resumen_imputacion_excel(
    wsi,
    filas: list[dict[str, Any]],
    *,
    titulo: str,
) -> None:
    negrita = Font(bold=True)
    t1 = wsi.cell(row=1, column=1, value=titulo)
    t1.font = negrita
    t1.alignment = Alignment(horizontal="center", vertical="center")
    wsi.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    wsi.cell(row=2, column=1, value="Cód. imputación").font = negrita
    wsi.cell(row=2, column=2, value="Imputación contable").font = negrita
    wsi.cell(row=2, column=3, value="Neto").font = negrita
    wsi.cell(row=2, column=4, value="IVA").font = negrita
    wsi.cell(row=2, column=5, value="Total").font = negrita
    fila = 3
    for row in filas:
        wsi.cell(row=fila, column=1, value=str(row.get("codigo_imputacion", "") or ""))
        wsi.cell(row=fila, column=2, value=str(row.get("nombre_imputacion", "") or ""))
        _celda_num(wsi, fila, 3, row.get("neto", 0))
        _celda_num(wsi, fila, 4, row.get("iva", 0))
        _celda_num(wsi, fila, 5, row.get("total", 0))
        fila += 1
    end_r = max(2, len(filas) + 2)
    wsi.auto_filter.ref = f"A2:E{end_r}"
    _ajustar_anchos_hoja_resumen_imputacion_excel(wsi)


def _hoja_comprobantes_desde_dataframe(
    wb,
    df_ajustado: pd.DataFrame,
    nombre_hoja: str,
    emitidos: bool,
) -> None:
    ws = wb.create_sheet(nombre_hoja)
    for row in dataframe_to_rows(df_ajustado, index=False, header=True):
        ws.append(row)
    encab = [c.value for c in ws[1]]
    _aplicar_hoja_comprobantes_excel(wb, ws, encab, emitidos, titulo_hoja=nombre_hoja)


def escribir_excel_informe_completo(
    df_ajustado: pd.DataFrame,
    destino: io.BytesIO | Path | str,
    *,
    emitidos: bool,
    totales: dict[str, float],
    totales_por_periodo: dict[str, dict[str, float]],
    periodos_orden: list[str],
    notas_credito_extras: dict,
    totales_resumen: dict,
    totales_detalle: dict,
    suma_total: float,
    columnas_orden: list[str],
    tabla_contrapartes: list[dict],
    resumen_imputacion: list[dict[str, Any]] | None = None,
    con_columnas_imputacion_en_contrapartes: bool = False,
) -> None:
    temp = io.BytesIO()
    df_ajustado.to_excel(temp, index=False, engine="openpyxl", sheet_name=_SHEET_COMPR)
    temp.seek(0)
    wb = load_workbook(temp)
    ws0 = wb.active
    encab = [c.value for c in ws0[1]]
    _aplicar_hoja_comprobantes_excel(wb, ws0, encab, emitidos)

    wsr = wb.create_sheet("Resumen", 1)
    _rellenar_hoja_resumen_excel(
        wsr, totales_resumen, totales_detalle, suma_total, notas_credito_extras
    )

    wsd = wb.create_sheet("Distribución mensual", 2)
    _rellenar_hoja_distribucion_excel(
        wsd, periodos_orden, columnas_orden, totales_por_periodo, notas_credito_extras
    )

    nom_total = "Total clientes" if emitidos else "Total proveedores"
    wsp = wb.create_sheet(nom_total, 3)
    _rellenar_hoja_contrapartes_excel(
        wsp,
        tabla_contrapartes,
        con_imputacion=con_columnas_imputacion_en_contrapartes,
    )

    if resumen_imputacion is not None:
        tit_imp = (
            "Resumen por imputación (clientes)"
            if emitidos
            else "Resumen por imputación (proveedores)"
        )
        wsi = wb.create_sheet("Resumen imputación", 4)
        _rellenar_hoja_resumen_imputacion_excel(
            wsi, resumen_imputacion, titulo=tit_imp
        )

    if isinstance(destino, io.BytesIO):
        destino.seek(0)
        destino.truncate(0)
        wb.save(destino)
        destino.seek(0)
    else:
        wb.save(destino)


def escribir_excel_informe_dual(
    destino: io.BytesIO | Path | str,
    *,
    df_recibidos: pd.DataFrame,
    totales_por_periodo_rec: dict[str, dict[str, float]],
    periodos_orden_rec: list[str],
    notas_credito_extras_rec: dict,
    totales_resumen_rec: dict,
    totales_detalle_rec: dict,
    suma_total_rec: float,
    tabla_contrapartes_rec: list[dict],
    df_emitidos: pd.DataFrame,
    totales_por_periodo_emit: dict[str, dict[str, float]],
    periodos_orden_emit: list[str],
    notas_credito_extras_emit: dict,
    totales_resumen_emit: dict,
    totales_detalle_emit: dict,
    suma_total_emit: float,
    tabla_contrapartes_emit: list[dict],
    columnas_orden: list[str],
    resumen_imputacion_rec: list[dict[str, Any]] | None = None,
    resumen_imputacion_emit: list[dict[str, Any]] | None = None,
    con_columnas_imputacion_en_contrapartes: bool = False,
) -> None:
    """Un solo libro con comprobantes, resumen, distribución y contrapartes para recibidos y emitidos."""
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Comprobantes recibidos"
    for row in dataframe_to_rows(df_recibidos, index=False, header=True):
        ws0.append(row)
    encab_r = [c.value for c in ws0[1]]
    _aplicar_hoja_comprobantes_excel(
        wb, ws0, encab_r, False, titulo_hoja="Comprobantes recibidos"
    )

    wsr = wb.create_sheet("Resumen recibidos")
    _rellenar_hoja_resumen_excel(
        wsr,
        totales_resumen_rec,
        totales_detalle_rec,
        suma_total_rec,
        notas_credito_extras_rec,
    )

    wsd_r = wb.create_sheet("Distribución mensual Recibidos")
    _rellenar_hoja_distribucion_excel(
        wsd_r,
        periodos_orden_rec,
        columnas_orden,
        totales_por_periodo_rec,
        notas_credito_extras_rec,
    )

    wsp_r = wb.create_sheet("Proveedores")
    _rellenar_hoja_contrapartes_excel(
        wsp_r,
        tabla_contrapartes_rec,
        con_imputacion=con_columnas_imputacion_en_contrapartes,
    )

    if resumen_imputacion_rec is not None:
        wsi_r = wb.create_sheet("Resumen imputación Recibidos")
        _rellenar_hoja_resumen_imputacion_excel(
            wsi_r,
            resumen_imputacion_rec,
            titulo="Resumen por imputación (proveedores)",
        )

    _hoja_comprobantes_desde_dataframe(
        wb, df_emitidos, "Comprobantes emitidos", True
    )

    wse = wb.create_sheet("Resumen emitidos")
    _rellenar_hoja_resumen_excel(
        wse,
        totales_resumen_emit,
        totales_detalle_emit,
        suma_total_emit,
        notas_credito_extras_emit,
    )

    wsd_e = wb.create_sheet("Distribución mensual Emitidos")
    _rellenar_hoja_distribucion_excel(
        wsd_e,
        periodos_orden_emit,
        columnas_orden,
        totales_por_periodo_emit,
        notas_credito_extras_emit,
    )

    wsp_e = wb.create_sheet("Clientes")
    _rellenar_hoja_contrapartes_excel(
        wsp_e,
        tabla_contrapartes_emit,
        con_imputacion=con_columnas_imputacion_en_contrapartes,
    )

    if resumen_imputacion_emit is not None:
        wsi_e = wb.create_sheet("Resumen imputación Emitidos")
        _rellenar_hoja_resumen_imputacion_excel(
            wsi_e,
            resumen_imputacion_emit,
            titulo="Resumen por imputación (clientes)",
        )

    if isinstance(destino, io.BytesIO):
        destino.seek(0)
        destino.truncate(0)
        wb.save(destino)
        destino.seek(0)
    else:
        wb.save(destino)


def escribir_excel_ajustado_con_formato(
    df: pd.DataFrame, destino: io.BytesIO | Path | str
) -> None:
    """
    Escribe solo la hoja de comprobantes con formato (compat. con llamadas
    que no construyen hojas de resumen / distribución).
    """
    temp = io.BytesIO()
    df.to_excel(temp, index=False, engine="openpyxl", sheet_name=_SHEET_COMPR)
    temp.seek(0)
    wb = load_workbook(temp)
    ws = wb.active
    encab = [c.value for c in ws[1]]
    _aplicar_hoja_comprobantes_excel(wb, ws, encab, emitidos=False)
    if isinstance(destino, io.BytesIO):
        destino.seek(0)
        destino.truncate(0)
        wb.save(destino)
        destino.seek(0)
    else:
        wb.save(destino)


def construir_ruta_salida(ruta_excel: str, salida_arg: str | None) -> Path:
    """Devuelve la ruta de salida para el excel ajustado."""
    if salida_arg:
        return Path(salida_arg)

    origen = Path(ruta_excel)
    return origen.with_name(f"{origen.stem}_ajustado.xlsx")


def main():
    if len(sys.argv) < 2:
        print("Uso: python sumar_imp_total.py <archivo.xlsx|archivo.csv> [hoja] [salida.xlsx]")
        print("  hoja: nombre o número de hoja (opcional, por defecto la primera)")
        print("  salida.xlsx: ruta del excel de salida (opcional)")
        print("Columnas sumadas desde la fila 3:", ", ".join(COLUMNAS_A_AJUSTAR))
        sys.exit(1)

    ruta = limpiar_argumento_ruta(sys.argv[1])
    hoja = sys.argv[2] if len(sys.argv) > 2 else 0
    salida_arg = limpiar_argumento_ruta(sys.argv[3]) if len(sys.argv) > 3 else None
    if isinstance(hoja, str) and hoja.isdigit():
        hoja = int(hoja)

    try:
        (
            df_ajustado,
            totales,
            tpp,
            nce,
            tabla_c,
        ) = procesar_archivo(ruta, hoja, nombre_archivo=ruta)
        ruta_salida = construir_ruta_salida(ruta, salida_arg)
        per_keys = periodos_orden_crono(
            tpp, nce.get("neto_nc_por_periodo", {}), nce.get("iva_nc_por_periodo", {})
        )
        totales_res = {c: totales[c] for c in COLUMNAS_TOTAL_RESUMEN}
        totales_det = {c: totales[c] for c in COLUMNAS_DETALLE_SIN_RESUMEN}
        escribir_excel_informe_completo(
            df_ajustado,
            ruta_salida,
            emitidos=False,
            totales=totales,
            totales_por_periodo=tpp,
            periodos_orden=per_keys,
            notas_credito_extras=nce,
            totales_resumen=totales_res,
            totales_detalle=totales_det,
            suma_total=total_resumen_pantalla(totales),
            columnas_orden=COLUMNAS_A_AJUSTAR,
            tabla_contrapartes=tabla_c,
        )

        print("Sumas (desde fila 3):")
        for col, total in totales.items():
            print(f"  {col}: {total:,.2f}")
        print("  ---")
        print(f"  Suma total (resumen): {total_resumen_pantalla(totales):,.2f}")
        print(f"Excel ajustado generado en: {ruta_salida}")
    except FileNotFoundError:
        print(f"Error: No se encontró el archivo '{ruta}'", file=sys.stderr)
        sys.exit(2)
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(3)


if __name__ == "__main__":
    main()
