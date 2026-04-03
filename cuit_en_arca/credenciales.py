"""Lectura del .xlsx de credenciales (celdas A1–C1 encabezados o A2–C2 datos)."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass

from openpyxl import load_workbook

from cuit_en_arca.errores import CredencialesArchivoError
from cuit_en_arca.validacion import parsear_rango_fechas_texto


@dataclass(frozen=True)
class CredencialesArca:
    cuit_login: str
    clave_fiscal: str
    cuit_representado: str
    """Si vienen de columna D (Rango Fechas), tienen prioridad sobre el formulario web."""
    rango_fecha_desde: str | None = None
    rango_fecha_hasta: str | None = None


def _solo_digitos(s: str, esperado: int = 11) -> str:
    d = re.sub(r"\D", "", str(s))
    if len(d) != esperado:
        raise CredencialesArchivoError(
            f"CUIT inválido: se esperaban {esperado} dígitos (valor recibido no normalizable)."
        )
    return d


def _parece_titulo_celda(val) -> bool:
    if val is None:
        return False
    t = str(val).strip().lower()
    if not t:
        return False
    return any(
        x in t
        for x in (
            "cuit",
            "clave",
            "fiscal",
            "represent",
            "representado",
            "representante",
        )
    )


def leer_credenciales_xlsx(buf: io.BytesIO) -> CredencialesArca:
    """
    Plantilla esperada:
    - Fila 1: títulos A1–C1 (CUIT representante, Clave fiscal, CUIT representado),
      **D1** = «Rango Fechas», valores en A2–D2 (en D2: ``dd/mm/yyyy - dd/mm/yyyy``); **o**
    - Una sola fila de valores en A1–D1 (sin títulos), siendo D1 el rango de fechas.
    """
    try:
        buf.seek(0)
        wb = load_workbook(buf, read_only=True, data_only=True)
        ws = wb.active
        a1 = ws["A1"].value
        b1 = ws["B1"].value
        c1 = ws["C1"].value
        a2 = ws["A2"].value
        b2 = ws["B2"].value
        c2 = ws["C2"].value
        d1 = ws["D1"].value
        d2 = ws["D2"].value
        wb.close()
    except Exception as exc:
        raise CredencialesArchivoError("No se pudo leer el archivo Excel de credenciales.") from exc

    titulos_abc = (
        _parece_titulo_celda(a1)
        or _parece_titulo_celda(b1)
        or _parece_titulo_celda(c1)
    )
    if titulos_abc:
        cuit_log, clave, cuit_repr = a2, b2, c2
        raw_rango = d2
    else:
        cuit_log, clave, cuit_repr = a1, b1, c1
        raw_rango = d1
        if raw_rango is None or str(raw_rango).strip() == "":
            raw_rango = d2

    rango_desde, rango_hasta = None, None
    par = parsear_rango_fechas_texto(raw_rango)
    if par:
        rango_desde, rango_hasta = par

    if clave is None or str(clave).strip() == "":
        raise CredencialesArchivoError("Falta la clave fiscal en el archivo (columna B).")

    if isinstance(cuit_log, float):
        cuit_log = f"{cuit_log:.0f}"
    if isinstance(cuit_repr, float):
        cuit_repr = f"{cuit_repr:.0f}"

    try:
        cuit_login = _solo_digitos(cuit_log)
        cuit_representado = _solo_digitos(cuit_repr)
    except CredencialesArchivoError:
        raise
    except Exception as exc:
        raise CredencialesArchivoError("Revisar CUITs en columnas A y C.") from exc

    return CredencialesArca(
        cuit_login=cuit_login,
        clave_fiscal=str(clave).strip(),
        cuit_representado=cuit_representado,
        rango_fecha_desde=rango_desde,
        rango_fecha_hasta=rango_hasta,
    )
